#!/usr/bin/env python3
"""Extract the compact 2024 bridge TSV from the harmonized bridge workbook."""
from __future__ import annotations
import argparse,csv
from datetime import date,datetime
from pathlib import Path
from openpyxl import load_workbook

def scalar(v):
    if v is None:return ''
    if isinstance(v,datetime):return v.date().isoformat()
    if isinstance(v,date):return v.isoformat()
    if isinstance(v,bool):return '1' if v else '0'
    if isinstance(v,float) and v.is_integer():return str(int(v))
    return str(v).replace('\t',' ').replace('\r',' ').replace('\n',' ')

def main():
    ap=argparse.ArgumentParser();ap.add_argument('workbook',type=Path);ap.add_argument('output',type=Path);a=ap.parse_args()
    wb=load_workbook(a.workbook,read_only=True,data_only=True)
    try:
        ws=wb['All Invoices']; rows=ws.iter_rows(values_only=True); headers=next(rows)
        m={str(v).strip():i for i,v in enumerate(headers) if v is not None}
        source=['Invoice ID','Source file','Invoice date','Due date','Invoice value','Client anonymised code','Supplier anonymised code','Series','Invoice number','Status']
        missing=[x for x in source if x not in m]
        if missing:raise ValueError(f'Missing columns: {missing}')
        a.output.parent.mkdir(parents=True,exist_ok=True)
        n=0
        with a.output.open('w',encoding='utf-8',newline='') as f:
            w=csv.writer(f,delimiter='\t',lineterminator='\r\n')
            w.writerow(['row_id','source_file','issue','due','value','buyer_code','supplier_code','series','number','invoice_status'])
            for row in rows:
                w.writerow([scalar(row[m[x]]) for x in source]);n+=1
        print(f'bridge_rows={n} output={a.output}')
    finally:wb.close()
if __name__=='__main__':main()
